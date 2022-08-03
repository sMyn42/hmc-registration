# Create Spreadsheet

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils import all_committees, global_caps, global_current, out_file

#main 

if __name__ == "__main__":
  if not os.path.exists(out_file):

    committee_assignments = Workbook()

    ws = committee_assignments.active
    ws.title = "Main"

    # Create column headers in Row 1:
    # House/Sen/Special

    d = [0 for i in all_committees]
    caps = [global_caps[c] for c in all_committees]
    currents = [global_current[c] for c in all_committees]
    left = []

    main_df = pd.DataFrame({"Committee" : all_committees, "Current" : currents, "Cap" : caps, "Left" : d})

    for r in dataframe_to_rows(main_df, index=True, header=True):
      ws.append(r)

    for cell in ws['A'] + ws[1]:
      cell.style = 'Pandas'
    
    for c in all_committees:
      committee_assignments.create_sheet(title=c)
    

    committee_assignments.save(out_file)
  else:
    print("Spreadsheet already exists! Delete before rerunning this script.")



# Add committee-specific spreadsheets

# columns include ID, First, Last, School, Email

# permit reference sheets to be created