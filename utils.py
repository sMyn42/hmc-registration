import numpy as np
import pandas as pd
from openpyxl import load_workbook
from itertools import islice
from openpyxl.utils.dataframe import dataframe_to_rows

out_file = "committee_assignments.xlsx"

house_committees = [
  'House Armed Services',
  'House Ed Labor (NOVICE)',
  'House Energy and Commerce',
  'House Financial Services',
  'House Foreign Affairs',
  'House Homeland Security',
  'House Judiciary',
  'House OGR',
  'House Intel',
  'House Climate Change',
  'House SST',
  'House Transpo and Infra',
  'House Vet Affairs'
]
senate_committees = [
  'Senate Armed Services',
  'Senate BHUA',
  'Senate Commerce',
  'Senate Energy NOVICE',
  'Senate EPW',
  'Senate Finance',
  'Senate Foreign Relations',
  'Senate HELP',
  'Senate Homeland Security',
  'Senate Judiciary NOVICE',
  'Senate Small Biz',
  'Senate Intel'
]
special_committees = [
  'G20',
  'HistComm',
  'Media',
  'NEC',
  'NSC',
  'PresCab',
  'SCOTUS',
  'West Wing',
  'WHO',
  'UNSC',
  'ConCon',
  'District Court'
]
team_committees = ["SCOTUS", "G20", "District Court"]

global_caps = {
  'House Armed Services' : 50,
  'House Ed Labor (NOVICE)' : 50,
  'House Energy and Commerce' : 50,
  'House Financial Services' : 50,
  'House Foreign Affairs' : 50,
  'House Homeland Security' : 50,
  'House Judiciary' : 50,
  'House OGR' : 50,
  'House Intel' : 50,
  'House Climate Change' : 50,
  'House SST' : 50,
  'House Transpo and Infra' : 50,
  'House Vet Affairs' : 50,
  'Senate Armed Services' : 50,
  'Senate BHUA' : 50,
  'Senate Commerce' : 50,
  'Senate Energy NOVICE' : 50,
  'Senate EPW' : 50,
  'Senate Finance' : 50,
  'Senate Foreign Relations' : 50,
  'Senate HELP' : 50,
  'Senate Homeland Security' : 50,
  'Senate Judiciary NOVICE' : 50,
  'Senate Small Biz' : 50,
  'Senate Intel' : 50,
  'G20' : 50,
  'HistComm' : 50,
  'Media' : 50,
  'NEC' : 50,
  'NSC' : 50,
  'PresCab' : 50,
  'SCOTUS' : 50,
  'West Wing' : 50,
  'WHO' : 50,
  'UNSC' : 50,
  'ConCon' : 50,
  'District Court' : 48 
}

all_committees = house_committees+senate_committees+special_committees
global_current = {c:0 for c in all_committees}

sheet_to_pref_dict = {'House Armed Services': 'House Armed Services',
 'House Ed Labor (NOVICE)': 'House Education and Labor',
 'House Energy and Commerce': 'House Energy and Commerce',
 'House Financial Services': 'House Financial Services',
 'House Foreign Affairs': 'House Foreign Affairs',
 'House Homeland Security': 'House Homeland Security',
 'House Judiciary': 'House Judiciary',
 'House OGR': 'House Oversight and Government Reform',
 'House Intel': 'House Select Committee on Intelligence',
 'House Climate Change': 'House Select Committee on Climate Crisis',
 'House SST': 'House Science Space and Technology',
 'House Transpo and Infra': 'House Transportation and Infrastructure',
 'House Vet Affairs': 'House Veterans Affairs',
 'Senate Armed Services': 'Senate Armed Services',
 'Senate BHUA': 'Senate Banking Housing and Urban Affairs',
 'Senate Commerce': 'Senate Commerce Science and Transportation',
 'Senate Energy NOVICE': 'Senate Energy and Natural Resources',
 'Senate EPW': 'Senate Environment and Public Works',
 'Senate Finance': 'Senate Finance',
 'Senate Foreign Relations': 'Senate Foreign Relations',
 'Senate HELP': 'Senate Health Education and Labor Pensions',
 'Senate Homeland Security': 'Senate Homeland Security and Government Affairs',
 'Senate Judiciary NOVICE': 'Senate Judiciary',
 'Senate Small Biz': 'Senate Small Business and Entrepreneurship',
 'Senate Intel': 'Senate Select Committee on Intelligence',
 'G20': 'Group of 20',
 'HistComm': 'Historical Committee',
 'Media': 'Media',
 'NEC': 'National Economic Council',
 'NSC': 'National Security Council',
 'PresCab': 'Presidential Cabinet',
 'SCOTUS': 'Supreme Court',
 'West Wing': 'West Wing',
 'WHO': 'World Health Organization',
 'UNSC': 'United Nations Security Council',
 'ConCon': 'Constitutional Convention',
 'District Court': 'District Court'}

pref_to_sheet_dict = dict([(v,k) for k, v in sheet_to_pref_dict.items()])

def get_assignment(pref_array, cap_dict): 

  # THERE IS DUPLICATE CODE

  # prefarray should be 8 entries long

  # also during team iteration
  for p in pref_array:
    p = p.replace("*", "")
    if p not in pref_to_sheet_dict.keys():
      print("Invalid Key Used in Role Request: " + str(p))
    elif p in team_committees:

      ##########
      # go to main flow control -> at end of each school, check team committees, and reassign to normal if teams are full. 
      # make sure caps (local, and global) are handled for teams. 
      ##########


      p = pref_to_sheet_dict[p]
      if cap_dict["local_current"][p] < cap_dict["local_caps"][p] and global_current[p] < global_caps[p]:
        return p
        # caps get updated outside of function
  for p in pref_array:
    p = p.replace("*", "")
    if p not in pref_to_sheet_dict.keys():
      print("Invalid Key Used in Role Request: " + str(p))
    else:
      p = pref_to_sheet_dict[p]
      if cap_dict["local_current"][p] < cap_dict["local_caps"][p] and global_current[p] < global_caps[p]:
        return p
  # nothing worked assign next available committee -> global caps can be changed using the spreadsheet itself.
  for comm in all_committees:
    # get local stat, global stat, compare with local cap, global cap, respectively
    # if the committee is valid, add person to it.
    if cap_dict["local_current"][comm] < cap_dict["local_caps"][comm] and global_current[comm] < global_caps[comm]:
      return comm
  max_comm = "NULL"
  max_stat = 0
  for comm, stat in global_current.items():
    if stat >= max_stat:
      max_comm = comm
      max_stat = stat
  return max_comm

def get_solo_assignment(pref_array, cap_dict):

  # TODO FIX DUPLICATE CODE HERE TOO!

  for p in pref_array:
    p = p.replace("*", "")
    if p not in pref_to_sheet_dict.keys():
      print("Invalid Key Used in Role Request: " + str(p))
    else:
      p = pref_to_sheet_dict[p]
      if cap_dict["local_current"][p] < cap_dict["local_caps"][p] and global_current[p] < global_caps[p]:
        return p
  # nothing worked assign next available committee -> global caps can be changed using the spreadsheet itself.
  for comm in all_committees:
    # get local stat, global stat, compare with local cap, global cap, respectively
    # if the committee is valid, add person to it.
    if cap_dict["local_current"][comm] < cap_dict["local_caps"][comm] and global_current[comm] < global_caps[comm]:
      return comm
  max_comm = "NULL"
  max_stat = 0
  for comm, stat in global_current.items():
    if stat >= max_stat:
      max_comm = comm
      max_stat = stat
  return max_comm

def update_cap_sheet(gc, ws):
  currents = [gc[c] for c in all_committees]
  data = ws.values
  cols = next(data)[1:]
  data = list(data)[1:]
  idx = [r[0] for r in data]
  data = (islice(r, 1, None) for r in data)
  df = pd.DataFrame(data, index=idx, columns=cols)
  df["Current"] = currents

  #Clear WS - think of a better way
  ws.delete_rows(1, ws.max_row)

  for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
  return ws

def import_workbook():
  committees = load_workbook(out_file)
  committees_raw = load_workbook("committee_assignments.xlsx", data_only=True)
  return (committees, committees_raw)

def read_global_current(ws):
  data = ws.values
  cols = next(data)[1:]
  data = list(data)[1:]
  idx = [r[0] for r in data]
  data = (islice(r, 1, None) for r in data)
  df = pd.DataFrame(data, index=idx, columns=cols)
  gc = dict(zip(df["Committee"], df["Current"]))
  return gc

